# -*- coding: utf-8 -*-
import re
import os
import sys
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from urllib.parse import quote
import json

if getattr(sys, "frozen", False):
    os.chdir(os.path.dirname(sys.executable))
else:
    os.chdir(os.path.dirname(os.path.abspath(__file__)))

LS_FILE = "livesklad.xlsx"
A1_FILE = "avito1.xlsx"
A2_FILE = "avito2.xlsx"
MAPPING_FILE = "mapping.xlsx"
OUT_FILE = "report.xlsx"

AVITO_ACTIVE = {"активно", "активное", "active"}
AVITO_NEW_STATES = {"новое", "новое с биркой"}
EXCLUDE_GROUP = "поступление"
EXCLUDE_NAME_KEYWORDS = ["чехол", "стекло", "защитн", "пленк", "запчаст"]
YES_WORDS = {"да", "1", "+", "yes", "ок", "ok"}
SOLD_WORDS = {"продано", "sold"}
SOLD_MARK = "SOLD"
SPLIT_WORD = "разделить"
UNMAP_MARK = "UNMAP"
UNMAP_WORDS = {"отвязать", "открепить", "-", "0"}

RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
LINK_FONT = Font(color="0563C1", underline="single")

AVITO2_SCHEME = ""
def detect_avito2_scheme():
    global AVITO2_SCHEME
    AVITO2_SCHEME = ""
    if os.name != "nt":
        return ""
    try:
        import winreg
        key = winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, r"avito2\shell\open\command", 0, winreg.KEY_READ)
        winreg.CloseKey(key)
        AVITO2_SCHEME = "avito2"
    except Exception:
        AVITO2_SCHEME = ""
    return AVITO2_SCHEME

def build_search_urls(av_id):
    filters = {
        "tabs": "active",
        "search": str(av_id),
        "daysOnAvito": {"from": None, "to": None},
        "price": {"from": None, "to": None},
        "statistics": {
            "fields": ["views", "contacts", "favorites", "viewsToContactsConversion", "calls"],
            "from": "2026-08-05",
            "to": "2026-09-03"
        }
    }
    filters_json = json.dumps(filters, separators=(',', ':'))
    encoded = quote(filters_json, safe='')
    https_url = f"https://www.avito.ru/profile/pro/items?filters={encoded}&pageFrom=1&pageTo=1&sortDirection=desc&sortField=sort_time"
    avito2_url = f"avito2://profile/pro/items?filters={encoded}&pageFrom=1&pageTo=1&sortDirection=desc&sortField=sort_time"
    return https_url, avito2_url

BRAND_CANON = {
    "apple":"apple","iphone":"apple","ipad":"apple","samsung":"samsung",
    "xiaomi":"xiaomi","redmi":"redmi","poco":"poco","honor":"honor","huawei":"huawei",
    "oppo":"oppo","vivo":"vivo","realme":"realme","oneplus":"oneplus","google":"google",
    "asus":"asus","lenovo":"lenovo","acer":"acer","dell":"dell","hp":"hp","msi":"msi",
    "sony":"sony","yandex":"yandex","яндекс":"yandex",
}
BRAND_VALUES = set(BRAND_CANON.values())
BRAND_WORDS = BRAND_VALUES | {"galaxy","смартфон","телефон","ноутбук","планшет","стилус","наушники","часы","приставки","приставка","консоль"}
DIFF_WORDS = {"pro","max","ultra","plus","lite","mini","neo","active","s","e","se","5g","air","anc","eco"}
PRODUCT_WORDS = {"геймпад","дисковод","станция","зарядная","контроллер","приставка","консоль","док"}
LAPTOP_MARKERS = ("macbook","magicbook","rog","zephyrus","thinkpad","ideapad","nitro","vivobook","zenbook","swift")
WATCH_MARKERS = ("watch","часы")
ROMAN = {"i":"1","ii":"2","iii":"3","iv":"4","v":"5","vi":"6","vii":"7","viii":"8","ix":"9","x":"10"}

LS_COLOR_CANON = {
    "black":"черный","space black":"черный","midnight":"черный","черный":"черный","чёрный":"черный","charcoal":"черный",
    "graphite":"серый","gray":"серый","grey":"серый","silver shadow":"серый",
    "white":"белый","starlight":"белый","белый":"белый",
    "silver":"серебристый","sliver":"серебристый","natural titanium":"серебристый","серебристый":"серебристый",
    "sterling silver":"серый+серебристый",
    "whitesilver":"белый+серебристый","silverblue":"синий",
    "gold":"золотистый","desert titanium":"золотистый","desert":"золотистый","sand":"бежевый",
    "золотой":"золотистый","золотистый":"золотистый","бежевый":"бежевый",
    "blue":"синий","navy":"синий","deep blue":"синий","indigo":"синий","blue titanium":"синий","ultramarine":"синий","синий":"синий",
    "cobalt blue":"синий","starlight blue":"голубой",
    "sky blue":"голубой","sky":"голубой","icyblue":"голубой","light blue":"голубой","misty blue":"голубой","mist blue":"голубой","teal":"зеленый","голубой":"голубой",
    "green":"зеленый","sage":"зеленый","зеленый":"зеленый","зелёный":"зеленый",
    "yellow":"желтый","желтый":"желтый","red":"красный","красный":"красный",
    "pink":"розовый","blush":"розовый","розовый":"розовый",
    "purple":"фиолетовый","lavender":"фиолетовый","lilac":"фиолетовый","cobalt violet":"фиолетовый","фиолетовый":"фиолетовый",
    "orange":"оранжевый","cosmic orange":"оранжевый","оранжевый":"оранжевый",
    "brown":"коричневый","коричневый":"коричневый",
}
AVITO_COLOR_CANON = {
    "черный":"черный","чёрный":"черный","белый":"белый",
    "серебристый":"серебристый","серебряный":"серебристый",
    "золотистый":"золотистый","золотой":"золотистый","бежевый":"бежевый",
    "серый":"серый","темно-серый":"серый","синий":"синий","голубой":"голубой",
    "зеленый":"зеленый","зелёный":"зеленый","желтый":"желтый","жёлтый":"желтый",
    "красный":"красный","бордовый":"бордовый","розовый":"розовый","фиолетовый":"фиолетовый",
    "оранжевый":"оранжевый","коричневый":"коричневый",
}
COLOR_SCAN = [
    ("silver shadow","серый"),("navy","синий"),
    ("silverblue","синий"),("whitesilver","белый+серебристый"),
    ("sterling silver","серый+серебристый"),
    ("cobalt blue","синий"),("starlight blue","голубой"),
    ("teal","зеленый"),
    ("icyblue","голубой"),("light blue","голубой"),("sand","бежевый"),("desert","золотистый"),("бежев","бежевый"),
    ("sky blue","голубой"),("mist blue","голубой"),
    ("син","синий"),("blue","синий"),("navy","синий"),("ultramarine","синий"),("indigo","синий"),
    ("голуб","голубой"),
    ("темно-сер","серый"),("grey","серый"),("gray","серый"),("graphite","серый"),
    ("чёрн","черный"),("черн","черный"),("black","черный"),("midnight","черный"),("charcoal","серый"),
    ("бел","белый"),("white","белый"),("starlight","белый"),
    ("серебр","серебристый"),("silver","серебристый"),("sliver","серебристый"),
    ("золот","золотистый"),("gold","золотистый"),
    ("зелён","зеленый"),("зелен","зеленый"),("green","зеленый"),("sage","зеленый"),
    ("желт","желтый"),("yellow","желтый"),
    ("красн","красный"),("red","красный"),("бордов","бордовый"),
    ("малинов","розовый"),("розов","розовый"),("pink","розовый"),("blush","розовый"),
    ("фиолет","фиолетовый"),("purple","фиолетовый"),("lavender","фиолетовый"),("lilac","фиолетовый"),
    ("оранжев","оранжевый"),("orange","оранжевый"),
    ("коричн","коричневый"),("brown","коричневый"),
]
COLOR_WORD_RE = re.compile(
    r"\b(черн\w*|чёрн\w*|бел\w*|серебр\w*|золот\w*|сер\w*|син\w*|голуб\w*|зелен\w*|зелён\w*|"
    r"красн\w*|розов\w*|малинов\w*|фиолет\w*|оранжев\w*|коричн\w*|желт\w*|"
    r"black|white|silver|sliver|gold|gray|grey|blue|green|red|pink|purple|orange|brown|yellow|"
    r"midnight|starlight|navy|teal|sage|lavender|lilac|indigo|charcoal|desert|natural|titanium|space|sky|graphite)\b", re.I)

# Регулярка для S/N - захватывает и латиницу, и кириллицу (для случаев типа С00STD...)
SN_RE = re.compile(r"s/n\s*[\(\[]?\s*([A-Za-zА-Яа-яЁё0-9][A-Za-zА-Яа-яЁё0-9-]*)\s*[\)\]]?", re.I)

def norm(s): return str(s).strip().lower() if pd.notna(s) else ""
def clean_code(c):
    c = str(c).strip()
    if re.fullmatch(r"\d+\.0", c): c = c[:-2]
    m = re.match(r"(\d+)", c)
    return m.group(1) if m and c else c
def scan_color(s):
    s = norm(s)
    for stem, canon in COLOR_SCAN:
        if stem in s: return canon
    return ""
def clean_color_frag(s):
    s = re.split(r"\s+-\s+", s)[0]
    s = re.sub(r"(sim\+?esim|esim|2\s*sim|asis\+?|imei.*|донор|до\s.*|не\s*вкл\.?|wi-?fi)", " ", s, flags=re.I)
    return re.sub(r"\s+", " ", s).strip()
def ls_color(frag, full_name):
    f = norm(clean_color_frag(frag))
    if f:
        if f in LS_COLOR_CANON: return LS_COLOR_CANON[f]
        f2 = f.replace("titanium","").replace("титан","").strip()
        if f2 in LS_COLOR_CANON: return LS_COLOR_CANON[f2]
        v = scan_color(f)
        if v: return v
    return scan_color(full_name)
def avito_color(raw, name):
    r = norm(raw)
    if r in AVITO_COLOR_CANON: return AVITO_COLOR_CANON[r]
    v = scan_color(raw)
    return v or scan_color(name)
def color_ok(c1, c2):
    if not c1 or not c2: return True
    s1 = set(c1.split("+")); s2 = set(c2.split("+"))
    if s1 & s2: return True
    def has(s,*c): return any(x in s for x in c)
    if (has(s1,"серый") and has(s2,"серебристый")) or (has(s2,"серый") and has(s1,"серебристый")): return True
    if (has(s1,"синий") and has(s2,"голубой")) or (has(s2,"синий") and has(s1,"голубой")): return True
    if (has(s1,"серый") and has(s2,"черный")) or (has(s2,"серый") and has(s1,"черный")): return True
    return False
def ad_eff_color(ad): return ad["color"] or ad.get("color2") or ""
def ad_color_match(color, ad):
    return color_ok(color, ad["color"]) or bool(ad.get("color2")) and color_ok(color, ad["color2"])
def is_excluded_name(s): return any(k in norm(s) for k in EXCLUDE_NAME_KEYWORDS)
def model_tokens(s): return set(t for t in re.findall(r"[a-zа-яё0-9]+", s) if len(t) > 1 or t.isdigit())
def digit_tokens(s): return set(t for t in model_tokens(s) if t.isdigit())
def jaccard(a, b):
    ta, tb = model_tokens(a), model_tokens(b)
    if not ta or not tb: return 0.0
    return len(ta & tb) / len(ta | tb)
def strip_brand(model):
    return " ".join(p for p in model.split(" ") if p not in BRAND_WORDS and not re.match(r"^\d+(gb|гб|tb|тб)$", p))
def split_brand(model):
    parts = model.split(" ")
    b = parts[0] if parts and parts[0] in BRAND_VALUES else ""
    r = " ".join(p for p in parts if p not in BRAND_WORDS)
    return b, r
def clean_cpu(rest):
    parts = rest.split(" "); out = []; skip = False
    for i, p in enumerate(parts):
        if skip: skip = False; continue
        is_chip = (re.fullmatch(r"i\d{1,2}[a-z]?", p) or re.fullmatch(r"[am]\d{1,2}", p)
                   or re.fullmatch(r"\d{2,5}[a-z]{1,2}", p) or re.fullmatch(r"[a-z]{2,4}\d{3,4}[a-z]?", p))
        if is_chip:
            if i+1 < len(parts) and parts[i+1] in ("pro","ultra"): skip = True
            continue
        out.append(p)
    return " ".join(out)

def clean_watch(rest):
    parts = rest.split(" ")
    out = []
    for p in parts:
        if re.fullmatch(r"\d+\s*mm", p): continue
        if p in ("gps","cellular","lte","wi-fi","wifi","bluetooth"): continue
        if p in ("aluminum","aluminium","stainless","steel","titanium"): continue
        out.append(p)
    return " ".join(out)

def cmp_rest(model):
    b, r = split_brand(model)
    if any(m in r for m in LAPTOP_MARKERS): r = clean_cpu(r)
    if any(m in r for m in WATCH_MARKERS): r = clean_watch(r)
    return b, r

def brands_compatible(b1, b2): return (not b1) or (not b2) or b1 == b2
def model_exact_ok(m1, m2):
    b1, r1 = cmp_rest(m1); b2, r2 = cmp_rest(m2)
    if not r1 or not r2: return m1 == m2
    return r1 == r2 and brands_compatible(b1, b2)
def _forbidden_diff(t1, t2):
    diff = t1 ^ t2
    if not diff: return False
    if diff & DIFF_WORDS: return True
    if diff & PRODUCT_WORDS: return True
    for t in diff:
        if t.isdigit(): return True
        if re.fullmatch(r"\d+(mm|мм)", t): return True
    return False
def fuzzy_model_ok(m1, m2):
    b1, r1 = cmp_rest(m1); b2, r2 = cmp_rest(m2)
    if not brands_compatible(b1, b2): return False
    t1, t2 = model_tokens(r1), model_tokens(r2)
    if not t1 or not t2: return False
    inter = len(t1 & t2)
    if inter < 2: return False
    if inter / len(t1 | t2) < 0.6: return False
    return not _forbidden_diff(t1, t2)
def to_num(s):
    if pd.isna(s): return None
    if isinstance(s, (int, float)): return float(s)
    t = re.sub(r"[^\d,.-]", "", str(s).replace("\xa0","").replace(" ",""))
    t = t.replace(",", ".")
    try: return float(t)
    except ValueError: return None
def norm_id(x):
    try: return str(int(x))
    except (ValueError, TypeError): return str(x).strip()
def strip_color_words(s): return re.sub(r"\s+", " ", COLOR_WORD_RE.sub(" ", s)).strip()
def norm_model(s):
    s = norm(s)
    s = s.replace("/"," ").replace("usb-c"," ").replace("type-c"," ").replace("wi-fi"," ").replace("wifi"," ")
    s = s.replace("ps5"," playstation 5").replace("ps4"," playstation 4").replace("+"," plus")
    s = re.sub(r"\bdual\s*sense\b", " dualsense ", s)
    s = s.replace("dualsense", " геймпад playstation 5 ")
    s = re.sub(r"\s+"," ",s).strip()
    if "яндекс" in s or "yandex" in s:
        # Обрабатываем все варианты Яндекс Станций, включая Дуо
        s = re.sub(r"\bяндекс\s+(станция\s+)?(лайт|стрит|макс|миди)\s*(дуо)?\b", r"яндекс станция \2 \3", s, flags=re.I)
        s = re.sub(r"\s+", " ", s)
        s = re.sub(r"\(?\s*с\s+часами\s*\)?", " withclock ", s, flags=re.I)
        s = re.sub(r"\(?\s*без\s+часов\s*\)?", " ", s, flags=re.I)
        s = re.sub(r"\(?\s*с\s+алисой\s+на\s+yagpt\s*\)?", " ", s, flags=re.I)
        s = re.sub(r"\(?\s*с\s+алисой\s*\)?", " ", s, flags=re.I)
        s = re.sub(r"\(?\s*на\s+yagpt\s*\)?", " ", s, flags=re.I)
        s = re.sub(r"\(?\s*на\s+алисе\s*\)?", " ", s, flags=re.I)
    s = re.sub(r"\b(asis\+?|as-is|as is|репак|repak|rep)\b", " ", s, flags=re.I)
    s = re.sub(r"\b(usb|lightning|для|стилус|алисой|алиса|yagpt)\b", " ", s)
    s = re.sub(r"\(\s*(\d+)\s*(st|nd|rd|th)\s*(gen|generation)\s*\)", r" \1 ", s, flags=re.I)
    s = re.sub(r"\b(\d+)\s*(st|nd|rd|th)\s*(gen|generation)\b", r" \1 ", s, flags=re.I)
    s = re.sub(r"\([^)]*\)", " ", s)
    s = strip_color_words(s)
    s = re.sub(r"\s+"," ",s).strip(" ,;-–/")
    parts = [ROMAN.get(p, p) for p in s.split(" ") if p]
    if "galaxy" in parts:
        parts = [p for p in parts if p != "galaxy"]
        if "samsung" not in parts: parts.insert(0, "samsung")
    if ("poco" in parts or "redmi" in parts) and "xiaomi" in parts:
        parts = [p for p in parts if p != "xiaomi"]
    if parts and parts[0] in BRAND_CANON: parts[0] = BRAND_CANON[parts[0]]
    return " ".join(parts)
def _mem_val(n, unit):
    n = int(n); return n*1024 if unit.lower() in ("tb","тб") else n
def extract_model_mem(clean):
    m = re.search(r"(\d+)\s*/\s*(\d+)\s*(gb|гб|tb|тб)", clean, re.I)
    if m: return clean[:m.start()].strip(" -–,"), int(m.group(1)), _mem_val(m.group(2), m.group(3)), clean[m.end():].strip(" -–,")
    m = re.search(r"(\d+)\s*(gb|гб|tb|тб)", clean, re.I)
    if m: return clean[:m.start()].strip(" -–,"), None, _mem_val(m.group(1), m.group(2)), clean[m.end():].strip(" -–,")
    return clean.strip(" -–,"), None, None, ""
def parse_ram(s):
    m = re.search(r"(\d+)\s*/\s*(\d+)\s*(gb|гб)", norm(s), re.I)
    return int(m.group(1)) if m else None
def model_from_name(name):
    model, ram, mem, _ = extract_model_mem(norm(name))
    return norm_model(model)
def norm_mem(s):
    m = re.search(r"(\d+)\s*/\s*(\d+)\s*(gb|гб|tb|тб)", norm(s), re.I)
    if m: return _mem_val(m.group(2), m.group(3))
    m = re.search(r"(\d+)\s*(gb|гб|tb|тб)", norm(s), re.I)
    return _mem_val(m.group(1), m.group(2)) if m else None
def norm_sim(s):
    s = norm(s)
    s = re.sub(r"(б/у|used|asis\+?|as-is|as is|active|перепак|nano)", "", s)
    s = s.replace(" ","").replace("-","")
    if "esim+esim" in s or "esimesim" in s: return "esim"
    if "esim" in s and "sim" in s.replace("esim",""): return "sim+esim"
    if ("2" in s or "dual" in s or "две" in s) and "sim" in s and "esim" not in s: return "2sim"
    if "esim" in s: return "esim"
    if "sim" in s: return "sim"
    return ""

def clean_ls_name(name):
    s = str(name)
    s = SN_RE.sub(" ", s)
    s = re.sub(r"imei\s*[\(\[]?\s*\d+\s*[\)\]]?", " ", s, flags=re.I)
    s = re.sub(r"\b(asis\+?|as-is|as is)\b", " ", s, flags=re.I)
    s = re.sub(r"\bб/у\b", " ", s, flags=re.I)
    s = re.sub(r"\s+", " ", s).strip(" -–,.")
    return s

def load_tables(path, required):
    sheets = pd.read_excel(path, sheet_name=None, header=None)
    frames = []
    for sh, df in sheets.items():
        if df is None or len(df) == 0: continue
        hdr_row = None
        for i in range(min(10, len(df))):
            cells = [norm(c) for c in df.iloc[i].tolist()]
            if all(any(r == c or (r and r in c) for c in cells) for r in required):
                hdr_row = i; break
        if hdr_row is None: continue
        seen = {}; head = []
        for c in df.iloc[hdr_row].tolist():
            h = norm(c) if pd.notna(c) else ""
            if h in seen: seen[h] += 1; h = f"{h}_{seen[h]}"
            else: seen[h] = 0
            head.append(h)
        body = df.iloc[hdr_row+1:].copy(); body.columns = head; body = body.dropna(how="all")
        frames.append(body)
    if not frames:
        raise ValueError(f"В файле {path} не найден лист с колонками {required}.\nЛисты: {list(sheets.keys())}")
    return pd.concat(frames, ignore_index=True)

def parse_ls_name(name):
    name = str(name)
    m2 = re.search(r"imei\s*\(?\s*(\d{13,17})\s*\)?", name, re.I)
    # Удаляем S/N (с поддержкой кириллицы) и IMEI из названия
    clean = SN_RE.sub(" ", name)
    if m2: clean = clean.replace(m2.group(0), " ")
    clean = re.sub(r"б/у", " ", clean, flags=re.I)
    clean = re.sub(r"\s+"," ",clean).strip(" -–")
    model, ram, mem, color_frag = extract_model_mem(clean)
    sim = ""
    sm = re.search(r"(esim\+esim|esim|2\s*sim|sim\+esim|sim)", clean, re.I)
    if sm: sim = norm_sim(sm.group(1))
    return model, ram, mem, color_frag, m2.group(1) if m2 else None, sim

def load_ls():
    df = load_tables(LS_FILE, ["наименование","код","розничная"])
    rows = []
    for _, r in df.iterrows():
        code = norm_id(r.get("код"))
        if not code or code == "nan" or not code.replace("_","").isalnum(): continue
        if is_excluded_name(r.get("наименование")): continue
        if (to_num(r.get("остаток")) or 0) <= 0: continue
        grp = norm(r.get("название группы")) + " " + norm(r.get("полная группа"))
        if EXCLUDE_GROUP in grp: continue
        model, ram, mem, color_frag, imei, sim_name = parse_ls_name(r["наименование"])
        sim_mod = norm_sim(r.get("модификация"))
        sim = sim_mod if sim_mod else sim_name
        color = ls_color(color_frag, r["наименование"])
        art = str(r.get("артикул")) if pd.notna(r.get("артикул")) else ""
        bat = None
        mb = re.search(r"(\d{2,3})\s*%", art)
        if mb: bat = int(mb.group(1))
        box = "box" in art.lower()
        used = ("б/у" in norm(r.get("название группы"))) or ("б/у" in norm(r["наименование"]))
        rows.append({"code":code,"name":str(r["наименование"]),"art":art,"model":norm_model(model),
                     "ram":ram,"mem":mem,"color":color,"sim":sim,"price":to_num(r["розничная"]),
                     "battery":bat,"box":box,"used":used})
    return rows

def find_color_column(df):
    for col in df.columns:
        cn = norm(col)
        if "цвет" in cn or "color" in cn: return col
    return None

def load_avito(path):
    df = load_tables(path, ["номер объявления на авито","цена"])
    rows = []; color_col = find_color_column(df)
    for _, r in df.iterrows():
        av_id = norm_id(r.get("номер объявления на авито"))
        if not av_id.isdigit(): continue
        st = norm(r.get("avitostatus"))
        if st and st not in AVITO_ACTIVE: continue
        name = str(r.get("название объявления"))
        if is_excluded_name(name): continue
        desc = str(r.get("описание объявления")) if pd.notna(r.get("описание объявления")) else ""
        raw_model = r.get("модель"); state = norm(r.get("состояние"))
        bat_v = to_num(r.get("состояние аккумулятора")); av_bat = int(bat_v) if bat_v is not None else None
        comp = norm(r.get("комплектация")); av_box = ("коробка" in comp) if comp else None
        model = norm_model(raw_model) if norm(raw_model) else model_from_name(name)
        mem = norm_mem(r.get("встроенная память"))
        if mem is None: mem = norm_mem(name)
        ram = parse_ram(name)
        if ram is None: ram = parse_ram(raw_model or "")
        sim = norm_sim(r.get("sim-карты"))
        if not sim: sim = norm_sim(name)
        color_raw = r.get(color_col) if color_col else None
        color = avito_color(color_raw, name); color2 = scan_color(name)
        is_new = state in AVITO_NEW_STATES
        rows.append({"id":av_id,"name":name,"desc":desc,"price":to_num(r["цена"]),"model":model,
                     "ram":ram,"mem":mem,"color":color,"color2":color2,"sim":sim,
                     "av_bat":av_bat,"av_box":av_box,"state":state,"is_new":is_new})
    return rows

def extract_ls_code(desc, code_set):
    if not desc: return None
    lines = [ln for ln in desc.split("\n") if ln.strip()]
    if not lines: return None
    last = re.sub(r"<[^>]+>", " ", lines[-1]); last = re.sub(r"\s+"," ",last).strip()
    for match in re.findall(r"\d+", last):
        if match in code_set: return match
    return None

def load_mapping():
    try:
        df = pd.read_excel(MAPPING_FILE)
        res = {}
        for _, r in df.iterrows():
            av_id = norm_id(r.get("avito_id")); ls_code = clean_code(r.get("ls_code"))
            if av_id and av_id != "nan" and ls_code: res[av_id] = ls_code
        return res
    except FileNotFoundError:
        return {}

MANUAL_SHEETS = ["Ручная проверка", "Ручная проверка (новые)", "Ручная проверка (БУ)"]
NOT_SHEETS = ["Нет в Avito1", "Нет в Avito2"]
FULL_SHEETS = ["Полный отчет (новые)", "Полный отчет БУ"]

def read_confirmations(ls_by_code, all_av_ids):
    extra = {}; split_pairs = {}
    try: names = pd.ExcelFile(OUT_FILE).sheet_names
    except Exception: return extra, split_pairs
    def parse_sheet(sheet):
        if sheet not in names: return
        try: d = pd.read_excel(OUT_FILE, sheet_name=sheet)
        except Exception: return
        has_accept = "Принять предложенный (да/нет)" in d.columns
        for _, r in d.iterrows():
            av1 = norm_id(r.get("Avito1 ID")) if "Avito1 ID" in r else ""
            av2 = norm_id(r.get("Avito2 ID")) if "Avito2 ID" in r else ""
            av0 = norm_id(r.get("Avito ID")) if "Avito ID" in r else ""
            targets = [a for a in (av1, av2, av0) if str(a).isdigit()]
            if not targets: continue
            usercode = clean_code(r.get("Ваш код LS")) if pd.notna(r.get("Ваш код LS")) else ""
            accept = norm(r.get("Принять предложенный (да/нет)")) if has_accept else ""
            if usercode.isdigit():
                for a in targets: extra[a] = usercode
            elif usercode in SOLD_WORDS:
                for a in targets: extra[a] = SOLD_MARK
            elif accept in YES_WORDS:
                p = clean_code(r.get("Предложенный код LS")) if pd.notna(r.get("Предложенный код LS")) else ""
                if p.isdigit():
                    for a in targets: extra[a] = p
            elif accept in SOLD_WORDS:
                for a in targets: extra[a] = SOLD_MARK
        return
    for sheet in MANUAL_SHEETS: parse_sheet(sheet)
    parse_sheet("Маппинг-черновик (новые)")
    parse_sheet("Маппинг-черновик (БУ)")
    parse_sheet("Проданы но активны")
    for sheet in NOT_SHEETS:
        if sheet not in names: continue
        try: d = pd.read_excel(OUT_FILE, sheet_name=sheet)
        except Exception: continue
        for _, r in d.iterrows():
            av = norm_id(r.get("Avito ID")) if "Avito ID" in r else ""
            if not str(av).isdigit(): continue
            code = clean_code(r.get("Ваш код LS")) if pd.notna(r.get("Ваш код LS")) else ""
            if not code: code = clean_code(r.get("Код LS")) if pd.notna(r.get("Код LS")) else ""
            if code and code not in SOLD_WORDS: extra[av] = code
    for sheet in FULL_SHEETS:
        if sheet not in names: continue
        try: d = pd.read_excel(OUT_FILE, sheet_name=sheet)
        except Exception: continue
        for _, r in d.iterrows():
            row_ls = clean_code(r.get("Код LS")) if "Код LS" in r else ""
            for idcol, corcol in (("Avito1 ID","Код A1 (правка)"), ("Avito2 ID","Код A2 (правка)")):
                av = norm_id(r.get(idcol)) if idcol in r else ""
                raw = r.get(corcol)
                if pd.isna(raw): continue
                V = clean_code(raw); val = norm(raw)
                if V in ls_by_code:
                    if str(av).isdigit(): extra[av] = V
                elif V.isdigit() and V in all_av_ids and row_ls in ls_by_code:
                    extra[V] = row_ls
                elif val in UNMAP_WORDS:
                    if str(av).isdigit(): extra[av] = UNMAP_MARK
    return extra, split_pairs

def save_mapping(manual, ls_by_code, av_name_by_id):
    rows = []
    for av_id, code in sorted(manual.items()):
        ls_item = ls_by_code.get(code)
        rows.append({"avito_id":av_id,"ls_code":code,
                     "ls_name":ls_item["name"] if ls_item else "",
                     "av_name":av_name_by_id.get(av_id, "")})
    pd.DataFrame(rows).to_excel(MAPPING_FILE, index=False)

def match_account(av, ls_rows, mapping, trusted_ids, unmap_ids):
    code_set = {r["code"] for r in ls_rows}
    ls_by_code = {r["code"]: r for r in ls_rows}
    code_to_ad = {}; sold_ads = []; used_ad_ids = set()
    
    for ad in av:
        if ad["id"] in unmap_ids: continue
        code = mapping.get(ad["id"])
        if code == UNMAP_MARK: continue
        if code == SOLD_MARK:
            sold_ads.append(ad); used_ad_ids.add(ad["id"]); continue
        if code and code not in ls_by_code and ad["id"] not in trusted_ids:
            sold_ads.append(ad); used_ad_ids.add(ad["id"]); continue
        
        desc_code = extract_ls_code(ad["desc"], code_set)
        
        if desc_code and desc_code in ls_by_code:
            if desc_code not in code_to_ad:
                code_to_ad[desc_code] = (ad["id"], ad["price"])
                used_ad_ids.add(ad["id"])
                continue
        
        if code and code in ls_by_code and ad["id"] not in trusted_ids:
            if desc_code != code:
                code = None
        
        if not code or code == "": code = desc_code
        if not code: continue
        if code == SOLD_MARK:
            sold_ads.append(ad); used_ad_ids.add(ad["id"]); continue
        if code in ls_by_code and code not in code_to_ad:
            code_to_ad[code] = (ad["id"], ad["price"]); used_ad_ids.add(ad["id"])
        elif code in ls_by_code and code in code_to_ad:
            pass
        else:
            sold_ads.append(ad); used_ad_ids.add(ad["id"])

    new_groups = {}; used_rows = []; code_to_key = {}
    for r in ls_rows:
        if r["used"]: used_rows.append(r)
        else:
            key = (r["model"], r["ram"], r["mem"], r["color"], r["sim"])
            new_groups.setdefault(key, []).append(r); code_to_key[r["code"]] = key

    group_ad = {}; used_ad = {}
    for code, (ad_id, price) in code_to_ad.items():
        ls_item = ls_by_code.get(code)
        if not ls_item: continue
        if ls_item["used"]: used_ad[code] = (ad_id, price)
        else:
            key = code_to_key.get(code)
            if key and key not in group_ad: group_ad[key] = (ad_id, price)

    rest_brands = {}
    for key in new_groups:
        b, r = cmp_rest(key[0]); rest_brands.setdefault(r, set()).add(b or "")

    def mem_ok(a, b):
        if a is None or b is None: return True
        return a == b

    def best_ad_for(model, color, sim, ram, mem, exact_model, exact_color):
        b_cur, r_cur = cmp_rest(model); best, best_j = None, -1.0
        for ad in av:
            if ad["id"] in used_ad_ids or ad["id"] in unmap_ids or not ad["is_new"]: continue
            if not mem_ok(ad["mem"], mem): continue
            ad_model = ad["model"]
            if exact_model:
                if not model_exact_ok(model, ad_model): continue
            else:
                if ad_model == model: continue
                ba, ra = cmp_rest(ad_model)
                if ra in rest_brands and (ba in rest_brands[ra] or "" in rest_brands[ra] or not ba) \
                   and not (ra == r_cur and brands_compatible(ba, b_cur)): continue
                if not fuzzy_model_ok(model, ad_model): continue
            if ram is not None and ad["ram"] is not None and ram != ad["ram"]: continue
            if sim and ad["sim"] and sim != ad["sim"]: continue
            if exact_color:
                if not (ad["color"] == color or ad.get("color2") == color): continue
            else:
                if not ad_color_match(color, ad): continue
            j2 = 1.0 if model_exact_ok(model, ad_model) else jaccard(r_cur, cmp_rest(ad_model)[1])
            if j2 > best_j: best_j, best = j2, ad
        return best

    for key, rows in new_groups.items():
        if key in group_ad: continue
        model, ram, mem, color, sim = key
        chosen = (best_ad_for(model,color,sim,ram,mem,True,True)
                  or best_ad_for(model,color,sim,ram,mem,True,False)
                  or best_ad_for(model,color,sim,ram,mem,False,True)
                  or best_ad_for(model,color,sim,ram,mem,False,False))
        if chosen:
            group_ad[key] = (chosen["id"], chosen["price"]); used_ad_ids.add(chosen["id"])

    unlinked = []
    for r in used_rows:
        if r["code"] in used_ad: continue
        best_score, best_ad = 0, None
        sb = strip_brand(r["model"])
        for ad in av:
            if ad["id"] in used_ad_ids or ad["id"] in unmap_ids or ad["is_new"] or ad["mem"] != r["mem"]: continue
            if not ad_color_match(r["color"], ad): continue
            jac = jaccard(sb, strip_brand(ad["model"]))
            if jac >= 0.85 and jac > best_score: best_score, best_ad = jac, ad
        if best_ad:
            used_ad[r["code"]] = (best_ad["id"], best_ad["price"]); used_ad_ids.add(best_ad["id"])
        else:
            unlinked.append(r)
    return group_ad, used_ad, sold_ads, unlinked, used_ad_ids

def best_ls_candidate(ad, ls_rows, matched_codes):
    ad_new = ad["is_new"]
    ba, ra = cmp_rest(ad["model"]); ad_digits = digit_tokens(ra)
    best, best_score = None, 0.0
    for r in ls_rows:
        if r["code"] in matched_codes: continue
        if r["used"] != (not ad_new): continue
        b, rr = cmp_rest(r["model"])
        if not brands_compatible(ba, b): continue
        t1, t2 = model_tokens(ra), model_tokens(rr)
        inter = len(t1 & t2)
        if inter < 2: continue
        if inter / len(t1 | t2) < 0.7: continue
        if digit_tokens(rr) != ad_digits: continue
        if _forbidden_diff(t1, t2): continue
        score = inter / len(t1 | t2)
        if ad_color_match(r["color"], ad): score += 0.2
        if score > best_score: best_score, best = score, r
    return best

def pair_accounts(unl1, unl2, split_pairs):
    used2 = set(); pairs = []
    for ad1 in unl1:
        b1, r1 = cmp_rest(ad1["model"]); best2, bestj = None, 0.0
        for ad2 in unl2:
            if ad2["id"] in used2 or ad2["is_new"] != ad1["is_new"]: continue
            if frozenset((ad1["id"], ad2["id"])) in split_pairs: continue
            b2, r2 = cmp_rest(ad2["model"])
            if not brands_compatible(b1, b2): continue
            t1, t2 = model_tokens(r1), model_tokens(r2)
            jac = jaccard(r1, r2)
            if jac < 0.6 or _forbidden_diff(t1, t2): continue
            if not color_ok(ad_eff_color(ad1), ad_eff_color(ad2)): continue
            if jac > bestj: bestj, best2 = jac, ad2
        if best2: used2.add(best2["id"])
        pairs.append((ad1, best2))
    for ad2 in unl2:
        if ad2["id"] not in used2:
            pairs.append((None, ad2))
    return pairs

def col_index(ws, name):
    for cell in ws[1]:
        if cell.value == name: return cell.column
    return None

RED_SHEETS = ["Полный отчет (новые)","Полный отчет БУ","Разница цен (новые)","Разница цен БУ"]
def highlight_and_link(wb):
    for sheet in wb.sheetnames:
        wb[sheet].freeze_panes = "A2"
    for sheet in RED_SHEETS:
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        cLS = col_index(ws, "Цена LS (мин)"); cA1 = col_index(ws, "Цена A1"); cA2 = col_index(ws, "Цена A2")
        if not cLS: continue
        for row in ws.iter_rows(min_row=2):
            ls = to_num(ws.cell(row=row[0].row, column=cLS).value); vals = []
            if cA1: vals.append(to_num(ws.cell(row=row[0].row, column=cA1).value))
            if cA2: vals.append(to_num(ws.cell(row=row[0].row, column=cA2).value))
            if ls is not None and any(v is not None and v != ls for v in vals):
                for cell in row: cell.fill = RED_FILL
    if "Сверка аккаунтов" in wb.sheetnames:
        ws = wb["Сверка аккаунтов"]
        for row in ws.iter_rows(min_row=2):
            for cell in row: cell.fill = YELLOW_FILL
    for sheet in ("Ручная проверка (новые)", "Ручная проверка (БУ)"):
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        cLS = col_index(ws, "Цена LS"); cP1 = col_index(ws, "Цена A1"); cP2 = col_index(ws, "Цена A2")
        for row in ws.iter_rows(min_row=2):
            pLS = to_num(ws.cell(row=row[0].row, column=cLS).value) if cLS else None
            p1 = to_num(ws.cell(row=row[0].row, column=cP1).value) if cP1 else None
            p2 = to_num(ws.cell(row=row[0].row, column=cP2).value) if cP2 else None
            fill = None
            if p1 is not None or p2 is not None:
                if pLS is not None and p1 is not None and p2 is not None and pLS == p1 == p2:
                    fill = GREEN_FILL
                elif pLS is not None and ((p1 is not None and p1 != pLS) or (p2 is not None and p2 != pLS)):
                    fill = RED_FILL
                elif p1 is not None and p2 is not None and p1 != p2:
                    fill = YELLOW_FILL
            if fill:
                for cell in row: cell.fill = fill
        opt_col = col_index(ws, "Принять предложенный (да/нет)")
        if opt_col:
            from openpyxl.worksheet.datavalidation import DataValidation
            dv = DataValidation(type="list", formula1='"Да,нет,продано,разделить"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{get_column_letter(opt_col)}2:{get_column_letter(opt_col)}5000")
    
    link_map = {
        "Полный отчет (новые)": [("Avito1 ID","default"),("Avito2 ID","chrome")],
        "Полный отчет БУ": [("Avito1 ID","default"),("Avito2 ID","chrome")],
        "Сверка аккаунтов": [("Avito1 ID","default"),("Avito2 ID","chrome")],
        "Ручная проверка (новые)": [("Avito1 ID","default"),("Avito2 ID","chrome")],
        "Ручная проверка (БУ)": [("Avito1 ID","default"),("Avito2 ID","chrome")],
        "Проданы но активны": [("Avito ID","default")],
        "Нет в Avito1": [("Avito ID","default")],
        "Нет в Avito2": [("Avito ID","chrome")],
    }
    for sheet, cols in link_map.items():
        if sheet not in wb.sheetnames: continue
        ws = wb[sheet]
        for name, mode in cols:
            c = col_index(ws, name)
            if not c: continue
            for row in ws.iter_rows(min_row=2, min_col=c, max_col=c):
                v = row[0].value
                if v in (None, "") or not str(v).isdigit(): continue
                https_url, avito2_url = build_search_urls(v)
                if mode == "chrome" and AVITO2_SCHEME:
                    row[0].hyperlink = avito2_url
                else:
                    row[0].hyperlink = https_url
                row[0].font = LINK_FONT

def main():
    detect_avito2_scheme()
    ls = load_ls(); a1 = load_avito(A1_FILE); a2 = load_avito(A2_FILE)
    total = len(ls)
    code_set = {r["code"] for r in ls}; ls_by_code = {r["code"]: r for r in ls}
    av_name_by_id = {ad["id"]: ad["name"] for ad in a1 + a2}
    all_av_ids = {ad["id"] for ad in a1} | {ad["id"] for ad in a2}

    base_mapping = load_mapping()
    extra, split_pairs = read_confirmations(ls_by_code, all_av_ids)
    base_mapping = {k:v for k,v in base_mapping.items() if k in all_av_ids}
    extra = {k:v for k,v in extra.items() if k in all_av_ids}
    trusted_ids = set(base_mapping.keys()) | set(extra.keys())

    mapping = {}
    for av_id, code in base_mapping.items(): mapping[av_id] = clean_code(code)
    for av_id, code in extra.items(): mapping[av_id] = clean_code(code)
    unmap_ids = {ad["id"] for ad in a1 + a2 if mapping.get(ad["id"]) == UNMAP_MARK}

    desc_code_count = 0
    for ad in a1 + a2:
        c = extract_ls_code(ad["desc"], code_set)
        if c:
            desc_code_count += 1
            if ad["id"] not in mapping: mapping[ad["id"]] = c

    new_confirms = {k: v for k, v in extra.items() if k not in base_mapping}
    manual_merged = dict(base_mapping); manual_merged.update(extra)
    save_mapping(manual_merged, ls_by_code, av_name_by_id)
    if new_confirms:
        print(f"Подтверждено привязок из report.xlsx: {len(new_confirms)} (сохранено в mapping.xlsx)")
    print(f"Объявлений с кодом LS в описании: {desc_code_count} | LS: {total} | Avito1: {len(a1)} | Avito2: {len(a2)}")

    g1, u1, sold1, unl1, busy1 = match_account(a1, ls, mapping, trusted_ids, unmap_ids)
    g2, u2, sold2, unl2, busy2 = match_account(a2, ls, mapping, trusted_ids, unmap_ids)

    new_groups = {}
    for r in ls:
        if not r["used"]:
            new_groups.setdefault((r["model"], r["ram"], r["mem"], r["color"], r["sim"]), []).append(r)

    matched_codes_a1 = set()
    matched_codes_a2 = set()
    
    for key, rows in new_groups.items():
        m1, m2 = g1.get(key), g2.get(key)
        for r in rows:
            if m1: matched_codes_a1.add(r["code"])
            if m2: matched_codes_a2.add(r["code"])
    
    for code in u1:
        matched_codes_a1.add(code)
    for code in u2:
        matched_codes_a2.add(code)
    
    matched_codes_either = matched_codes_a1 | matched_codes_a2
    
    matched_a1_count = len(matched_codes_a1)
    matched_a2_count = len(matched_codes_a2)
    matched_either_count = len(matched_codes_either)
    
    pct1 = matched_a1_count/total*100
    pct2 = matched_a2_count/total*100
    pct_either = matched_either_count/total*100

    full_new, full_used = [], []; diff_new, diff_used = [], []; not1, not2 = [], []
    for key, rows in new_groups.items():
        m1, m2 = g1.get(key), g2.get(key)
        exp = min(r["price"] for r in rows if r["price"] is not None) if any(r["price"] is not None for r in rows) else None
        if not m1: not1.append({"Тип":"новый","code":rows[0]["code"],"name":rows[0]["name"],"price":exp,"units":len(rows),
                                "Avito ID":"","Ваш код LS":rows[0]["code"]})
        if not m2: not2.append({"Тип":"новый","code":rows[0]["code"],"name":rows[0]["name"],"price":exp,"units":len(rows),
                                "Avito ID":"","Ваш код LS":rows[0]["code"]})
        d1 = (m1[1]-exp) if (m1 and m1[1] is not None and exp is not None) else None
        d2 = (m2[1]-exp) if (m2 and m2[1] is not None and exp is not None) else None
        d12 = (m1[1]-m2[1]) if (m1 and m2 and m1[1] is not None and m2[1] is not None) else None
        row = {"Код LS":rows[0]["code"]+(f" +{len(rows)-1}" if len(rows)>1 else ""),
               "Название":rows[0]["name"],
               "Название (чистое)":clean_ls_name(rows[0]["name"]),
               "Цена LS (мин)":exp,"Ед.":len(rows),
               "Avito1 ID":m1[0] if m1 else "","Цена A1":m1[1] if m1 else "",
               "Avito2 ID":m2[0] if m2 else "","Цена A2":m2[1] if m2 else "",
               "A1-LS":d1,"A2-LS":d2,"A1-A2":d12,
               "Код A1 (правка)":"","Код A2 (правка)":""}
        full_new.append(row)
        if d1 or d2 or d12: diff_new.append(row)

    for r in ls:
        if not r["used"]: continue
        m1, m2 = u1.get(r["code"]), u2.get(r["code"])
        if not m1: not1.append({"Тип":"б/у","code":r["code"],"name":r["name"],"price":r["price"],"units":1,
                                "Avito ID":"","Ваш код LS":r["code"]})
        if not m2: not2.append({"Тип":"б/у","code":r["code"],"name":r["name"],"price":r["price"],"units":1,
                                "Avito ID":"","Ваш код LS":r["code"]})
        d1 = (m1[1]-r["price"]) if (m1 and m1[1] is not None and r["price"] is not None) else None
        d2 = (m2[1]-r["price"]) if (m2 and m2[1] is not None and r["price"] is not None) else None
        d12 = (m1[1]-m2[1]) if (m1 and m2 and m1[1] is not None and m2[1] is not None) else None
        row = {"Код LS":r["code"],
               "Название":r["name"],
               "Название (чистое)":clean_ls_name(r["name"]),
               "Цена LS (мин)":r["price"],"Ед.":1,
               "Avito1 ID":m1[0] if m1 else "","Цена A1":m1[1] if m1 else "",
               "Avito2 ID":m2[0] if m2 else "","Цена A2":m2[1] if m2 else "",
               "A1-LS":d1,"A2-LS":d2,"A1-A2":d12,
               "Код A1 (правка)":"","Код A2 (правка)":""}
        full_used.append(row)
        if d1 or d2 or d12: diff_used.append(row)

    cross = []
    for typ, rows in (("новый", full_new), ("б/у", full_used)):
        for row in rows:
            p1, p2 = row["Цена A1"], row["Цена A2"]
            pr1 = p1 not in ("", None); pr2 = p2 not in ("", None)
            problem = ""
            if pr1 and pr2 and row["A1-A2"]: problem = "разные цены на аккаунтах"
            elif pr1 and not pr2: problem = "нет на Avito2"
            elif pr2 and not pr1: problem = "нет на Avito1"
            if problem:
                cross.append({"Тип":typ,"Проблема":problem,"Код LS":row["Код LS"],"Название":row["Название"],
                              "Цена LS (мин)":row["Цена LS (мин)"],"Avito1 ID":row["Avito1 ID"],"Цена A1":p1,
                              "Avito2 ID":row["Avito2 ID"],"Цена A2":p2,"A1-A2":row["A1-A2"]})

    def av_rows(lst, acc):
        return [{"Аккаунт":acc,"Avito ID":r["id"],"Название":r["name"],"Цена":r["price"],
                 "Модель":r["model"],"Память":r["mem"],"Цвет":r["color"],"Ваш код LS":""} for r in lst]

    matched1 = set(u1.keys()); matched2 = set(u2.keys())
    for key, rows in new_groups.items():
        if key in g1: matched1.add(rows[0]["code"])
        if key in g2: matched2.add(rows[0]["code"])

    unl1 = [ad for ad in a1 if ad["id"] not in busy1 and mapping.get(ad["id"]) != SOLD_MARK and mapping.get(ad["id"]) not in ls_by_code]
    unl2 = [ad for ad in a2 if ad["id"] not in busy2 and mapping.get(ad["id"]) != SOLD_MARK and mapping.get(ad["id"]) not in ls_by_code]
    pairs = pair_accounts(unl1, unl2, split_pairs)

    manual_new, manual_used = [], []
    for ad1, ad2 in pairs:
        ref = ad1 or ad2
        cand = best_ls_candidate(ref, ls, matched1 | matched2)
        row = {
            "Avito1 ID": ad1["id"] if ad1 else "",
            "Название A1": ad1["name"] if ad1 else "",
            "Цена A1": ad1["price"] if ad1 else "",
            "Цвет A1": ad_eff_color(ad1) if ad1 else "",
            "Avito2 ID": ad2["id"] if ad2 else "",
            "Название A2": ad2["name"] if ad2 else "",
            "Цена A2": ad2["price"] if ad2 else "",
            "Цвет A2": ad_eff_color(ad2) if ad2 else "",
            "Предложенный код LS": cand["code"] if cand else "",
            "Предложенное название LS": cand["name"] if cand else "",
            "Цена LS": cand["price"] if cand else "",
            "Принять предложенный (да/нет)": "",
            "Ваш код LS": "",
        }
        (manual_new if ref["is_new"] else manual_used).append(row)
    manual_new.sort(key=lambda r: (str(r["Название A1"] or r["Название A2"])))
    manual_used.sort(key=lambda r: (str(r["Название A1"] or r["Название A2"])))

    unlinked1 = len(unl1); unlinked2 = len(unl2)

    summary = [
        ["Позиций в LS", total],
        ["Новых групп в LS", len(new_groups)],
        ["Б/У позиций в LS", sum(1 for r in ls if r["used"])],
        ["Активных объявлений Avito1", len(a1)],
        ["Активных объявлений Avito2", len(a2)],
        ["Совпало Avito1 (новые)", len(g1)],
        ["Совпало Avito2 (новые)", len(g2)],
        ["Совпало Avito1 (б/у)", len(u1)],
        ["Совпало Avito2 (б/у)", len(u2)],
        ["Опубликовано из LS на Avito1 (позиций)", matched_a1_count],
        ["Опубликовано из LS на Avito2 (позиций)", matched_a2_count],
        ["Опубликовано из LS хотя бы на одном (позиций)", matched_either_count],
        ["Процент опубликованных из LS (Avito1)", f"{pct1:.1f}%"],
        ["Процент опубликованных из LS (Avito2)", f"{pct2:.1f}%"],
        ["Процент позиций LS на Авито (хотя бы один аккаунт)", f"{pct_either:.1f}%"],
        ["Нет в Avito1", len(not1)],
        ["Нет в Avito2", len(not2)],
        ["Непривязанных объявлений Avito1", unlinked1],
        ["Непривязанных объявлений Avito2", unlinked2],
        ["Проданы, но активны Avito1", len(sold1)],
        ["Проданы, но активны Avito2", len(sold2)],
        ["Расхождений между аккаунтами", len(cross)],
    ]

    with pd.ExcelWriter(OUT_FILE) as w:
        pd.DataFrame(summary, columns=["Показатель","Значение"]).to_excel(w, sheet_name="Сводка", index=False)
        pd.DataFrame(full_new).to_excel(w, sheet_name="Полный отчет (новые)", index=False)
        pd.DataFrame(full_used).to_excel(w, sheet_name="Полный отчет БУ", index=False)
        pd.DataFrame(diff_new).to_excel(w, sheet_name="Разница цен (новые)", index=False)
        pd.DataFrame(diff_used).to_excel(w, sheet_name="Разница цен БУ", index=False)
        pd.DataFrame(cross).to_excel(w, sheet_name="Сверка аккаунтов", index=False)
        pd.DataFrame(not1).to_excel(w, sheet_name="Нет в Avito1", index=False)
        pd.DataFrame(not2).to_excel(w, sheet_name="Нет в Avito2", index=False)
        sold_rows = av_rows(sold1,1)+av_rows(sold2,2)
        pd.DataFrame(sold_rows, columns=["Аккаунт","Avito ID","Название","Цена","Модель","Память","Цвет","Ваш код LS"]).to_excel(w, sheet_name="Проданы но активны", index=False)
        if manual_new:
            pd.DataFrame(manual_new).to_excel(w, sheet_name="Ручная проверка (новые)", index=False)
        if manual_used:
            pd.DataFrame(manual_used).to_excel(w, sheet_name="Ручная проверка (БУ)", index=False)
        highlight_and_link(w.book)

    print(f"\nИТОГ: Avito1 = {matched_a1_count}/{total} ({pct1:.1f}%) | Avito2 = {matched_a2_count}/{total} ({pct2:.1f}%) | Позиций LS на Авито: {matched_either_count}/{total} ({pct_either:.1f}%)")
    print("Готово:", OUT_FILE)
    for k, v in summary:
        print(f"{k}: {v}")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("ОШИБКА:", e)
    try:
        if os.name == "nt":
            os.system("pause")
        else:
            input("\nДля продолжения нажмите Enter . . .")
    except Exception:
        pass